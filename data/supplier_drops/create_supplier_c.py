"""
Creates supplier_c.xlsx — an Excel file with merged header cells
This simulates a real-world supplier who formats their Excel file with merged cells
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

output_path = "/Users/vishal/Documents/Internship Project/data/supplier_drops/supplier_c.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Supply Data"

# Add a merged title row (like a real supplier would do)
ws.merge_cells("A1:F1")
ws["A1"] = "HomeGoods Supplier — Monthly Supply Report — January 2026"
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = Alignment(horizontal="center")

# Merge category header cells (another quirk)
ws.merge_cells("A2:B2")
ws["A2"] = "PRODUCT INFO"
ws["A2"].font = Font(bold=True)
ws["A2"].alignment = Alignment(horizontal="center")

ws.merge_cells("C2:D2")
ws["C2"] = "PRICING"
ws["C2"].font = Font(bold=True)
ws["C2"].alignment = Alignment(horizontal="center")

ws.merge_cells("E2:F2")
ws["E2"] = "ORDER DETAILS"
ws["E2"].font = Font(bold=True)
ws["E2"].alignment = Alignment(horizontal="center")

# Actual column headers on row 3
headers = ["product_name", "category", "price", "quantity", "order_date", "supplier"]
for col, header in enumerate(headers, 1):
    ws.cell(row=3, column=col, value=header).font = Font(bold=True)

# Data rows starting from row 4
data = [
    ("Bed Sheet Set", "Bedding", 2500, 30, "2026-01-15", "HomeGoods"),
    ("Pillow Cover", "Bedding", 350, 80, "2026-01-15", "HomeGoods"),
    ("Bath Towel", "Bathroom", 600, 50, "2026-01-16", "HomeGoods"),
    ("Hand Towel", "Bathroom", 250, 100, "2026-01-16", "HomeGoods"),
    ("Kitchen Towel", "Kitchen", 180, 120, "2026-01-17", "HomeGoods"),
    ("Curtain Set", "Decor", 1800, 20, "2026-01-17", "HomeGoods"),
    ("Cushion Cover", "Decor", 450, 60, "2026-01-18", "HomeGoods"),
    ("Table Runner", "Dining", 380, 40, "2026-01-18", "HomeGoods"),
    ("Dinner Set", "Dining", 3200, 15, "2026-01-19", "HomeGoods"),
    ("Glass Set 6pc", "Dining", 850, 25, "2026-01-19", "HomeGoods"),
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Set column widths
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 18

wb.save(output_path)
print(f"✅ supplier_c.xlsx created at: {output_path}")
